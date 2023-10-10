import openpyxl
wb = openpyxl.load_workbook('Итоги3.xlsx')
l1 = wb['Лист1']

dct={}
for i in range(l1.max_row):
    dct[l1.cell(row=i + 1, column=1).value] = l1.cell(row=i + 1, column=2).value

print(dct)

lst=[]
for i,j in dct.items():
    lst.append(j)

min_z = min(lst)
max_z = max(lst)

k=0
summa=0
for i in lst:
    summa = summa + i
    k +=1
ar_m = round(summa/k,1)


s_lst=sorted(lst)
print(s_lst)

mediana = 0
if len(lst)%2 !=0:
    mediana = s_lst[len(lst)//2]
else:
    mediana = (s_lst[len(lst)//2-1] + s_lst[len(lst)//2])/2

l2 = wb.create_sheet("Info")

p1 = l2['A1'].value = 'Минимальное значение'
z1 = l2['B1'].value = min_z
p2 = l2['A2'].value = 'Максимальное значение'
z2 = l2['B2'].value = max_z
p3 = l2['A3'].value = 'Среднее арифметическое'
z3 = l2['B3'].value = ar_m
p4 = l2['A4'].value = 'Медиана'
z4 = l2['B4'].value = mediana

wb.save('Итоги3.xlsx')



