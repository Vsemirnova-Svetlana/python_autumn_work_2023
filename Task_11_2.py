import openpyxl
from openpyxl import Workbook

with open('Info.txt', encoding = 'utf-8') as f:
    k = f.readlines()

# print(k)
lst=[]
for i in k:
    lst.append(i.split(','))
# print(lst)

sorted_lst = sorted(lst, key = lambda x: (x[3],x[1],x[2]))
# print(sorted_lst)

wb = Workbook('Info')
wb.create_sheet("Data")


wb.save('Info.xlsx')

wb = openpyxl.load_workbook('Info.xlsx')
print(wb.sheetnames)
ws = wb["Data"]

itogo = 0
for k,i in enumerate(sorted_lst):
    ws[f"A{k+1}"] = i[0]
    ws[f'B{k+1}'] = i[1]
    ws[f"C{k+1}"] = i[2]
    ws[f"D{k+1}"] = i[3]
    ws[f"E{k+1}"] = int(i[4])
    itogo +=int(i[4])

f = ws.cell(row = ws.max_row + 1, column = ws.max_column - 1).value = 'ИТОГО'
p = ws.cell(row = ws.max_row,column = ws.max_column).value = itogo
wb.save('Info.xlsx')