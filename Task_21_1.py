import openpyxl
from openpyxl import Workbook
wb = Workbook()
wb.save('Разработчик на Питоне.xlsx')

wb = openpyxl.load_workbook('Разработчик на Питоне.xlsx')
wb.create_sheet('Pupil')
ws = wb['Pupil']
ws['A1'].value = '№'
ws['B1'].value = 'ФИО студента'
ws['C1'].value = 'Посещаемость, %'
ws['D1'].value = 'Выполнено задач, %'
ws['E1'].value = 'Получили диплом, чел.'
ws['F1'].value = 'Нашли работу,чел.'

wp=wb.create_sheet('Посещаемость')
wp['A1'].value = '№'
wp['B1'].value = 'ФИО/Дата'

wu = wb.create_sheet('Успеваемость') # указывается кол-во правильно выполненных задач (1-3)
wu['A1'].value ='№'
wu['B1'].value ='ФИО/Дата'

wd = wb.create_sheet('Дипломная работа')
wd['A1'].value = '№'
wd['B1'].value = 'ФИО'
wd['C1'].value = 'Название проекта'
wd['D1'].value = 'Оценка'

wo = wb.create_sheet('Оплата курса')
wo['A1'].value = '№'
wo['B1'].value = 'ФИО'
wo['C1'].value = 'статус оплаты 1 часть'
wo['D1'].value = 'статус оплаты 2 часть'

wb.save('Разработчик на Питоне.xlsx')
