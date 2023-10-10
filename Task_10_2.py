import openpyxl
wb = openpyxl.load_workbook("Итоги2.xlsx")
l1 = wb['Period1']
l2 = wb['Period2']
dct1={}
dct2={}

for i in range(l1.max_row):
    dct1[l1.cell(row=i + 1, column=1).value] = dct1.get(l1.cell(row=i + 1, column=1).value, 0) + l1.cell(row=i + 1,column = 2).value
print(dct1)

for i in range(l2.max_row):
    dct2[l2.cell(row=i+1,column = 1).value] = dct2.get(l2.cell(row = i+1, column = 1).value, 0 ) + l2.cell(row = i+1, column = 2).value
print(dct2)

wb.create_sheet("Сумма")
l3=wb["Сумма"]

for i, j in dct2.items():
    dct1[i] = dct1.get(i,0) + j
print(dct1)
total = 0
for i in dct1.values():
    total = total + i

d = sorted(dct1.items())
print(d)

k=1
for x in d:

    l3.cell(row=k, column=1).value = x[0]
    l3.cell(row=k, column=2).value = x[1]
    k = k+1


p = l3.cell(row=len(dct1)+1, column = 1).value = 'ИТОГО:'
v= l3.cell(row=len(dct1)+1, column = 2).value = total

wb.save("Итоги2.xlsx")