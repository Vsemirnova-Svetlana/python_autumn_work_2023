import openpyxl

wb = openpyxl.load_workbook("Итоги.xlsx")
l1 = wb['Лист1']

dct = {}
total = 0
for i in range(l1.max_row):
    dct[l1.cell(row=i+1, column=1).value] = dct.get(l1.cell(row=i+1, column=1).value,0)+l1.cell(row=i + 1, column=2).value
    total = total + l1.cell(row=i + 1, column=2).value

print(dct)
print(total)

l2 = wb['Лист2']

k=1
for i,j in dct.items():
    l2.cell(row=k, column=1).value = i
    l2.cell(row=k, column=2).value = j
    k = k+1
p = l2.cell(row=len(dct)+1, column = 1).value = 'ИТОГО:'
v= l2.cell(row=len(dct)+1, column = 2).value = total

wb.save("Итоги.xlsx")
